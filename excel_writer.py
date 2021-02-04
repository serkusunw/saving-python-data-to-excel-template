from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.formula.translate import Translator

excel_data: [[any]] = [
    ['A1', 'B1', 3, 5],
    ['A2', 'B2', 4, 33],
    ['A3', 'B3', 222, 23],
    ['A4', 'B4', 7, 4],
    ['A5', 'B5', 35, 62]
]
file_path: str = 'src/template.xlsx'

wb: Workbook = load_workbook(file_path)
ws: Worksheet = wb['sheet']
initial_row_number = ws.max_row
for row in excel_data:
    ws.append(row)
    copy_formula_to = 'E' + str(ws.max_row)
    translated_formula = Translator(ws['E3'].value, 'E3').translate_formula(copy_formula_to)
    ws[copy_formula_to] = translated_formula
ws.move_range('A4:E8', rows=-1, translate=True)
wb.save(file_path)
wb.close()
