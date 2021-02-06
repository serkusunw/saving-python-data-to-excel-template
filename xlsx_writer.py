from typing import Tuple, Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import TYPE_FORMULA, Cell
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from model.cell_details import CellDetails
from model.coordinate import Coordinate


class XlsxWriter:
    def __init__(self, file_path: str) -> None:
        self.file_path: str = file_path
        self._wb: Workbook = load_workbook(file_path)

    def write_rows_to_sheet(self, sheet_name: str, collection: [[any]], initial_row: int = 1,
                            initial_row_as_pattern: bool = False):
        ws: Worksheet = self._wb[sheet_name]
        # for conditional_formatting in ws.conditional_formatting._cf_rules:
        #     for cell_range in conditional_formatting.cells.ranges:
        #         print('cell contains a conditional formatting')
        pattern_row_formulas = self._get_row_formulas(ws, initial_row)
        for row in collection:
            ws.append(row)
            self._translate_formulas_to_row(ws, ws.max_row, pattern_row_formulas)
        ws.move_range('A4:G8', rows=-1, translate=True)

    @staticmethod
    def _translate_formulas_to_row(ws: Worksheet, row: int, cell_details_list: List[CellDetails]):
        for cell_details in cell_details_list:
            copy_formula_to = get_column_letter(cell_details.coordinate.col) + str(row)
            translator = Translator(cell_details.value, cell_details.coordinate.coordinate)
            translated_formula = translator.translate_formula(copy_formula_to)
            ws[copy_formula_to] = translated_formula
            ws[copy_formula_to].number_format = cell_details.cell_format

    @staticmethod
    def _get_row_formulas(ws: Worksheet, row: int, min_col: int = 0, max_col: int = 0) -> List[CellDetails]:
        row_formulas: List[CellDetails] = []
        first_row_index = 0
        row: Tuple[Cell] = list(ws.iter_rows(row, row, min_col, max_col))[first_row_index]
        for cell in row:
            if cell.data_type == TYPE_FORMULA:
                coordinate = Coordinate(cell.coordinate, cell.row, cell.column)
                cell_details = CellDetails(coordinate, cell.number_format, cell.value)
                row_formulas.append(cell_details)
        return row_formulas

    def save(self):
        self._wb.save(self.file_path)

    def close(self):
        self._wb.close()
